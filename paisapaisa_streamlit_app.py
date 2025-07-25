import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import tempfile

def generate_flowchart(df, output_path):
    df = df.rename(columns={
        'Account No./ (Wallet /PG/PA) Id': 'Sender',
        'Account No': 'Receiver',
        'Transaction Amount': 'Amount',
        'Bank/FIs': 'Bank',
        'Ifsc Code': 'IFSC'
    })

    df['Amount'] = pd.to_numeric(
        df['Amount'].astype(str).str.replace(",", "").str.replace("₹", ""),
        errors='coerce'
    )
    df = df[df['Amount'] > 50000]

    victim = df['Sender'].value_counts().idxmax()
    layer1_df = df[(df['Sender'] == victim) & (df['Receiver'].notna())]
    unique_l1 = layer1_df['Receiver'].unique().tolist()

    wb = Workbook()
    ws = wb.active
    ws.title = "Flowchart"

    # Styles
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_blue = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_violet = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    colspan = len(unique_l1) * 2 - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(colspan, 1))
    vcell = ws.cell(row=1, column=1, value=f"Victim: {victim}")
    vcell.fill = fill_blue
    vcell.alignment = align_center
    vcell.font = Font(bold=True)

    col = 1
    for l1 in unique_l1:
        row = 3
        ws.cell(row=row, column=col, value="↓").alignment = align_center
        row += 1

        l1_txn = df[(df['Sender'] == victim) & (df['Receiver'] == l1)].iloc[0]
        l1_text = f"Layer 1 Account\nBank: {l1_txn['Bank']}\nA/c No: {l1}\nIFSC: {l1_txn['IFSC']}\nSent: ₹{int(l1_txn['Amount']):,}"
        cell = ws.cell(row=row, column=col, value=l1_text)
        cell.fill = fill_green
        cell.alignment = align_center
        row += 2

        ws.cell(row=row, column=col, value="↓").alignment = align_center
        row += 1

        # Withdrawals from Layer 1
        l1_withdrawals = df[(df['Sender'] == l1) & (df['Receiver'].isna())]
        for _, wd in l1_withdrawals.iterrows():
            amt = wd['Amount']
            text = f"💸 Withdrawal Made\nFrom: Layer 1\nA/c No: {l1}\nAmount: ₹{int(amt):,}"
            cell = ws.cell(row=row, column=col, value=text)
            cell.fill = fill_yellow if amt <= 100000 else fill_red
            cell.alignment = align_center
            row += 2

        # Layer 2 Accounts
        l2_df = df[(df['Sender'] == l1) & (df['Receiver'].notna())]
        for _, l2_row in l2_df.iterrows():
            l2 = l2_row['Receiver']
            if pd.isna(l2):
                continue

            l2_text = f"Layer 2 Account\nBank: {l2_row['Bank']}\nA/c No: {l2}\nIFSC: {l2_row['IFSC']}\nReceived: ₹{int(l2_row['Amount']):,}"
            cell = ws.cell(row=row, column=col, value=l2_text)
            cell.fill = fill_violet
            cell.alignment = align_center
            row += 2

            ws.cell(row=row, column=col, value="↓").alignment = align_center
            row += 1

            # Withdrawals from Layer 2
            wd_df = df[(df['Sender'] == l2) & (df['Receiver'].isna())]
            for _, wd in wd_df.iterrows():
                amt = wd['Amount']
                text = f"💸 Withdrawal Made\nFrom: Layer 2\nA/c No: {l2}\nAmount: ₹{int(amt):,}"
                cell = ws.cell(row=row, column=col, value=text)
                cell.fill = fill_yellow if amt <= 100000 else fill_red
                cell.alignment = align_center
                row += 2

        col += 2

    for i in range(1, col):
        ws.column_dimensions[get_column_letter(i)].width = 28

    wb.save(output_path)

# ------------------------ Streamlit UI ----------------------------

st.set_page_config(page_title="Paisa Paisa 💸", layout="centered")

# Diwali lights background CSS
st.markdown("""
<style>
body {
    background-color: #0d0d0d;
    background-image:
        radial-gradient(circle at 10% 20%, #ffcc00 3px, transparent 3px),
        radial-gradient(circle at 20% 80%, #ff9900 3px, transparent 3px),
        radial-gradient(circle at 90% 25%, #ffaa33 3px, transparent 3px),
        radial-gradient(circle at 70% 60%, #ffcc66 3px, transparent 3px),
        radial-gradient(circle at 50% 90%, #ffd700 3px, transparent 3px),
        radial-gradient(circle at 80% 10%, #ffae42 3px, transparent 3px);
    background-size: 100% 100%;
    animation: twinkle 4s infinite ease-in-out;
    color: white;
}
@keyframes twinkle {
  0% { opacity: 0.95; }
  50% { opacity: 0.7; }
  100% { opacity: 0.95; }
}
</style>
""", unsafe_allow_html=True)

# App UI
st.title("🪔 Paisa Paisa Flowchart Generator")
st.markdown("Upload your Excel transaction sheet to visualize money trail layers and withdrawals with full styling.")

uploaded_file = st.file_uploader("📤 Upload Excel File", type=["xlsx"])
if uploaded_file:
    df = pd.read_excel(uploaded_file)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name
    generate_flowchart(df, output_path)
    st.success("✅ Flowchart Generated Successfully!")
    with open(output_path, "rb") as f:
        st.download_button("📥 Download Output Excel", f, file_name=uploaded_file.name.replace(".xlsx", "_output.xlsx"))
